import os
import glob
import json
from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string, coordinate_from_string

package_dir = os.path.dirname(os.path.abspath(__file__))
excel_files = glob.glob(f"{package_dir}/excel_files/*.xlsx")

SHT_IDX = 0
main_dict = []

def cell_coordinate(string):
    '''Get the coordiante of the row[0] and column[1]'''
    xy = coordinate_from_string(string)
    x = xy[1]
    y = column_index_from_string(xy[0])
    return {'x': x, 'y': y}

'''
This function is to loop through the well data parameters in the excel sheet
then it load it to a dictionary
'''

def get_parameters_mod(first_cell , last_cell, sht_idx):
    first_coor = cell_coordinate(first_cell)
    last_coor = cell_coordinate(last_cell)
    wellParamters = {}
    # for row in range(start_row, end_row):
    for row in range(first_coor['x'], last_coor['x']):
        for col in range(first_coor['y'], first_coor['y'] + 1):
            c_title = wb[wb.sheetnames[sht_idx]].cell(row, col)
            c_value = wb[wb.sheetnames[sht_idx]].cell(row, col+5)
            # add values to the dictionary using the key and value
            # if c_value is None:
            if c_value.value is None:
                continue
            wellParamters[c_title.value] = c_value.value
    return wellParamters


'''
This scrip is used to get the activities report and append it to a list than
can be used later to be added to a final report.
'''
# def get_activities(start_row, end_row, col, sht_idx,is_last:bool):
def get_activities_mod(first_cell , last_cell, sht_idx, is_last:bool):
    first_coor = cell_coordinate(first_cell)
    last_coor = cell_coordinate(last_cell)
    activity_list = []
    for row in range(first_coor['x'], last_coor['x']):
        for col in range(first_coor['y'], first_coor['y']+1):
            c_lastActivities = wb[wb.sheetnames[sht_idx]].cell(row, col)
            if c_lastActivities.value is None:
                continue
            if is_last:
                activity_list.append(c_lastActivities.value)
            else:
                activity_list.append(c_lastActivities.value)
    return activity_list


for excel_file in excel_files:
    wb = load_workbook(filename=excel_file, read_only=True)
    date = wb.worksheets[SHT_IDX]["I5"].value
    client = wb.worksheets[SHT_IDX]["F5"].value
    well = wb.worksheets[SHT_IDX]["F7"].value
    jobID = wb.worksheets[SHT_IDX]["I7"].value
    # get_parameters(22, 38, 9, 0) # old style
    parameters = get_parameters_mod('I18', 'I38', SHT_IDX)
    if client == 'DNO':
        last_activity = get_activities_mod('A49', 'A56', 0, True)
        next_activity = get_activities_mod('A57', 'A64', 0, False)
    elif client == 'HKN':
        last_activity = get_activities_mod('A68', 'A77', 0, True)
        next_activity = get_activities_mod('A79', 'A88', 0, False)
    elif client == 'TAQA':
        last_activity = get_activities_mod('A60', 'A66', 0, True)
        next_activity = get_activities_mod('A70', 'A77', 0, False)
    else:
        last_activity = get_activities_mod('A62', 'A71', 0, True)
        next_activity = get_activities_mod('A72', 'A78', 0, False)

    finalReport = {
        "client": client,
        "well": well,
        "jobID": jobID,
        "date": date,
        "Well Parameters": parameters,
        "last 24 activities": last_activity,
        "next 24 activities": next_activity,
        "file name": excel_file,
    }
    main_dict.append(finalReport)

with open('file_test.json', 'a') as f:
    json.dump(main_dict, f, default=str)


'''
for excel_file in excel_files:
    wb = load_workbook(filename=excel_file, read_only=True)

    supervisorName = wb.worksheets[0]["C6"].value
    date = wb.worksheets[0]["I5"].value
    client = wb.worksheets[0]["F5"].value
    well = wb.worksheets[0]["F7"].value
    jobID = wb.worksheets[0]["I7"].value
    FlowingHour = wb.worksheets[0]["N23"].value
    MaxOilRate = wb.worksheets[0]["N24"].value
    avgOilRate = wb.worksheets[0]["N25"].value
    MaxWaterRate = wb.worksheets[0]["N26"].value
    avgWaterRate = wb.worksheets[0]["N27"].value
    avgGasRate = wb.worksheets[0]["N28"].value
    staticPressure = wb.worksheets[0]["N29"].value
    diffPressure = wb.worksheets[0]["N30"].value
    CMF = wb.worksheets[0]["N31"].value
    H2S = wb.worksheets[0]["N33"].value
    CO2 = wb.worksheets[0]["N34"].value
    BSW = wb.worksheets[0]["N35"].value
    WHP = wb.worksheets[0]["N36"].value
    WHT = wb.worksheets[0]["N37"].value
    choke = wb.worksheets[0]["N38"].value

    dict_values = {
        "supervisor": supervisorName, "date": date, "client": client, "well": well, "jobID": jobID,
        "FlowingHour": FlowingHour, "MaxOilRate": MaxOilRate, "avgOilRate": avgOilRate,
        "MaxWaterRate": MaxWaterRate, "avgWaterRate": avgWaterRate, "avgGasRate": avgGasRate,
        "staticPressure": staticPressure, "diffPressure": diffPressure, "CMF": CMF, "H2S": H2S,
        "CO2": CO2, "BSW": BSW, "WHP": WHP, "WHT": WHT, "choke": choke, "fileName": excel_file
    }

    with open('db.csv', 'a') as csv_file:
        writer = csv.writer(csv_file)
        tmpList = []
        for key in dict_values.keys():
            tmpList.append(dict_values[key])
        writer.writerow(tmpList)
        print(tmpList)
'''

# Get the well activities
# for row in range(71,75):
#     for col in range(1,2):
#         c_nextActivity = wb[wb.sheetnames[0]].cell(row,col)
#         # add values to the list
#         nextActivity.append(c_nextActivity.value)

# Get the well activities
# for row in range(60,69):
#     for col in range(1,2):
#         c_lastActivities = wb[wb.sheetnames[0]].cell(row,col)
#         lastActivity.append(c_lastActivities.value)

# Get the well parameter values and titles
# for row in range(18,34):
#     for col in range(9,10):
#         c_title = wb[wb.sheetnames[0]].cell(row,col)
#         c_value = wb[wb.sheetnames[0]].cell(row,col+5)
#         # add values to the dictionary using the key and value
#         wellParamters[c_title.value] = c_value.value

