from openpyxl import load_workbook, Workbook  # pip install openpyxl
# from openpyxl.utils import get_column_letter, get_column_interval

'''
This code is directed to open last month crew time sheet and picks up the time sheet from 26 to 31 of last month and
past it in 4 sheets fom the ops stat excel sheet last it opens the current month excel and copy time sheet from 1st to
25th of the current month and then paste it to the ops stat sheet in the correct sheet and position
The code is very dirty and repetitive and is not factored but it do the task faster than I normally do
- I download the files from OnDrive and rename it to lastMonth and currentMonth and keep a copy of the opsstat sheet
in the same directory.
The code then creates a new file called test.xlsx having the values
Created on : 29 Sep 2021

To run this scrip and see how much it took use the code below
{date && python3 opsStat.py && date && open test.xlsx}
'''

lastMonth = "/Users/mohammedalbatati/Downloads/opsStat/lastMonth.xlsx"
currentMonth = "/Users/mohammedalbatati/Downloads/opsStat/currentMonth.xlsx"
ops_file = "/Users/mohammedalbatati/Downloads/opsStat/opsStat.xlsx"

wbl = load_workbook(lastMonth, read_only=True)
wbc = load_workbook(currentMonth, read_only=True)
wbo = load_workbook(ops_file)

'''
This function is to clear the sheet and make it empty
'''
def clearSheet(sheetname, therange):
    for row in wbo[sheetname][therange]:
        for cell in row:
            cell.value = None

'''
This function is the core of the code and it uses some parameters to get the job done
rs row start, re row end, cs column start, ce column end, ar adjust row, ac adjust column, source sheet name. destination sheet name, is current boolen
'''
def copyFromCurrent(rs, re, cs, ce, ar, ac, source, destination, is_current:bool):
    for row in range(rs, re):
        for col in range(cs, ce):
            if is_current is True:
                c = wbc[source].cell(row, col)
            else:
                c = wbl[source].cell(row, col)
            wbo[destination].cell(row+ar, col+ac).value = c.value

# Clear the sheet to be ready for data
clearSheet('Consultants', 'I13:AM136')
clearSheet('Emp.SWT', 'I13:AM136')
clearSheet('Emp.SLS', 'I13:AM136')
clearSheet('WTC OverHead', 'I13:AM28')

copyFromCurrent(3, 28, 30, 55, 10, -15, 'WTC', 'Emp.SWT',is_current = True)
copyFromCurrent(3, 28, 55, 61, 10, -46, 'WTC', 'Emp.SWT',is_current = False)
copyFromCurrent(28, 39, 30, 55, -15, -15, 'WTC', 'Emp.SLS',is_current = True)
copyFromCurrent(28, 39, 55, 61, -15, -46, 'WTC', 'Emp.SLS',is_current = False)
copyFromCurrent(3, 55, 30, 55, 10, -15, 'Consultnat', 'Consultants',is_current = True)
copyFromCurrent(3, 55, 55, 61, 10, -46, 'Consultnat', 'Consultants',is_current = False)
copyFromCurrent(14, 19, 5, 30, -1, 10, 'Timesheet', 'WTC OverHead',is_current = True)
copyFromCurrent(14, 19, 30 , 36, -1, -21, 'Timesheet', 'WTC OverHead',is_current = False)
wbo.save("test.xlsx")



'''
Below is the old attempt to make the code the first time and it was dirty code

# ========Getting Data from other sheets =========

# source WTC-SWT// dimension is from AD3 (row = 3 column =30 ) to BB28 (Row=28 colum=54)
# destination EMP.SWT is from O13 (row = 13 column =15 ) to AM46 (row = 46 and column 39)
# For SWT
wbl = load_workbook(lastMonth, read_only=True)
copyFromCurrent(3, 28, 30, 55, 10, -15, 'WTC', 'Emp.SWT',is_current = True)
# for row in range(3, 28):
#     for col in range(30, 55):
#         c = wbc['WTC'].cell(row, col)
#         wbo['Emp.SWT'].cell(row+10, col-15).value = c.value
# Last month
# source WTC-SWT// dimension is from BC3 (row = 3 column =55 ) to BH28 (Row=28 colum=60)
# destination EMP.SWT is from I13 (row = 13 column =9 ) 



copyFromCurrent(3, 28, 55, 61, 10, -46, 'WTC', 'Emp.SWT',is_current = False)
# for row in range(3, 28):
#     for col in range(55,61):
#         c = wbl['WTC'].cell(row, col)
#         wbo['Emp.SWT'].cell(row+10, col-46).value = c.value

# =================================
# source WTC-SLS// dimension is from AD8 (row = 28 column =30 ) to BB38 (Row=38 colum=54)
# destination EMP.SLS is from O13 (row = 13 column =15 ) to AM46 (row = 46 and column 39)
#For SLS
copyFromCurrent(28, 39, 30, 55, -15, -15, 'WTC', 'Emp.SLS',is_current = True)
# for row in range(28,39):
#     for col in range(30, 55):
#         c = wbc['WTC'].cell(row, col)
#         wbo['Emp.SLS'].cell(row-15, col-15).value = c.value

#last month
# destination EMP.SLS is from I13 (row = 13 column =9 ) to AM46 (row = 46 and column 39)
#For SLS
copyFromCurrent(28, 39, 55, 61, -15, -46, 'WTC', 'Emp.SLS',is_current = False)
# for row in range(28,39):
#     for col in range(55, 61):
#         c = wbl['WTC'].cell(row, col)
#         wbo['Emp.SLS'].cell(row-15, col-46).value = c.value

# =================================
# source Consultant// dimension is from AD3 (row = 3 column =30 ) to BB54 (Row=54 colum=54)
# destination Consultant is from O13 (row = 13 column =15 ) to AM134 (row = 134 and column 39)

#For Consultants
copyFromCurrent(3, 55, 30, 55, 10, -15, 'Consultnat', 'Consultants',is_current = True)
# for row in range(3, 55):
#     for col in range(30, 55):
#         # c = wbc['Consultant'].cell(row, col)
#         c = wbc['Consultnat'].cell(row, col)
#         wbo['Consultants'].cell(row+10, col-15).value = c.value

#Last month
#For Consultants
copyFromCurrent(3, 55, 55, 61, 10, -46, 'Consultnat', 'Consultants',is_current = False)
# for row in range(3, 55):
#     for col in range(55, 61):
#         c = wbl['Consultnat'].cell(row, col)
#         wbo['Consultants'].cell(row+10, col-46).value = c.value

# =================================
# source Timesheet// dimension is from E14 (row = 14 column =5 ) to AC18 (Row=18 colum=29)
# destination Timeshee is from O13 (row = 13 column =15 ) to AM23 (row = 23 and column 39)
#For Overhead
copyFromCurrent(14, 19, 5, 30, -1, 10, 'Timesheet', 'WTC OverHead',is_current = True)
# for row in range(14, 19):
#     # for col in range(5, 55):
#     for col in range(5, 30):
#         c = wbc['Timesheet'].cell(row, col)
#         wbo['WTC OverHead'].cell(row-1, col+10).value = c.value
# Last month
copyFromCurrent(14, 19, 30 , 36, -1, -21, 'Timesheet', 'WTC OverHead',is_current = False)
# for row in range(14, 19):
#     for col in range(30, 36):
#         c = wbl['Timesheet'].cell(row, col)
#         wbo['WTC OverHead'].cell(row-1, col-21).value = c.value

# Saving all the changes to a new file
wbo.save("test.xlsx")




'''
